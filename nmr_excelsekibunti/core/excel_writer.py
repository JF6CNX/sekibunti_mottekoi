import re
import os
from collections import defaultdict

from openpyxl import Workbook, load_workbook


# ==========================================
# 時間並び替え
# ==========================================
def parse_time_value(label):

    m = re.search(r"(\d+)(min|h|d)", label)

    if not m:
        return 999999

    value = int(m.group(1))
    unit = m.group(2)

    if unit == "min":
        return value / 60

    if unit == "h":
        return value

    if unit == "d":
        return value * 24

    return 999999


# ==========================================
# 積分データ取得
# ==========================================
def collect_integral_data(base_dirs):

    data = defaultdict(lambda: defaultdict(dict))
    ppm_range = defaultdict(dict)

    for base_dir in base_dirs:

        for root, dirs, files in os.walk(base_dir):

            for file in files:

                if "integrals" not in file.lower():
                    continue

                path = os.path.join(root, file)

                # --------------------------
                # サンプル名
                # --------------------------
                folder_name = os.path.basename(base_dir)

                # TTH-E01-001_1h
                # → sample=TTH-E01-001
                # → time=1h

                m = re.match(r"(.+?)_(.+)", folder_name)

                if m:
                    sample = m.group(1)
                    time_label = m.group(2)
                else:
                    sample = folder_name
                    time_label = "unknown"

                # --------------------------
                # integrals.txt 読み込み
                # --------------------------
                try:

                    with open(path, encoding="utf-8", errors="ignore") as f:

                        for line in f:

                            parts = line.split()

                            if len(parts) < 4:
                                continue

                            try:

                                number = int(parts[0])

                                start = float(parts[1])
                                end = float(parts[2])

                                integral = float(parts[3])

                                data[sample][time_label][number] = integral

                                if number not in ppm_range[sample]:
                                    ppm_range[sample][number] = (start, end)

                            except:
                                pass

                except:
                    pass

    return data, ppm_range


# ==========================================
# Excel出力
# ==========================================
def write_excel_from_integrals_multi(
    base_dirs,
    output_path,
    number_order=None,
    append=False,
):

    data, ppm_range = collect_integral_data(base_dirs)

    # ======================================
    # Workbook
    # ======================================
    if append and os.path.exists(output_path):

        wb = load_workbook(output_path)

    else:

        wb = Workbook()

        if wb.active:
            wb.remove(wb.active)

    # ======================================
    # サンプルごと
    # ======================================
    for sample, time_dict in data.items():

        # ------------------------------
        # number一覧
        # ------------------------------
        all_numbers = set()

        for t in time_dict.values():
            all_numbers.update(t.keys())

        if number_order:

            number_list = [
                n for n in number_order
                if n in all_numbers
            ]

            remain = sorted(all_numbers - set(number_list))

            number_list.extend(remain)

        else:

            number_list = sorted(all_numbers)

        # ------------------------------
        # sheet
        # ------------------------------
        sheet_name = sample[:31]

        if sheet_name in wb.sheetnames:
            del wb[sheet_name]

        ws = wb.create_sheet(title=sheet_name)

        # ------------------------------
        # header
        # ------------------------------
        ws.append(["time"] + number_list)

        ppm_row = ["ppm"]

        for n in number_list:

            if n in ppm_range[sample]:

                s, e = ppm_range[sample][n]

                ppm_row.append(f"{s:.2f}–{e:.2f}")

            else:

                ppm_row.append("")

        ws.append(ppm_row)

        # ------------------------------
        # data rows
        # ------------------------------
        sorted_times = sorted(
            time_dict.keys(),
            key=parse_time_value
        )

        for time_label in sorted_times:

            row = [time_label]

            for n in number_list:

                row.append(
                    time_dict[time_label].get(n, "")
                )

            ws.append(row)

    # ======================================
    # save
    # ======================================
    wb.save(output_path)