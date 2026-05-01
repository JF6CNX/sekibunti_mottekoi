import os
from openpyxl import load_workbook, Workbook
from pathlib import Path


def reorder_excel_auto(input_path):

    print("=== start ===")
    input_path = Path(input_path)

    if not input_path.exists():
        print("file not found")
        return

    custom_order = [1, 6, 4, 5, 3, 2]

    output_path = input_path.with_name(
        input_path.stem + "_reordered" + input_path.suffix
    )

    wb_in = load_workbook(input_path)
    wb_out = Workbook()
    wb_out.remove(wb_out.active)

    for sheet_name in wb_in.sheetnames:

        ws_in = wb_in[sheet_name]
        ws_out = wb_out.create_sheet(title=sheet_name[:31])

        rows = list(ws_in.iter_rows(values_only=True))
        if len(rows) < 2:
            continue

        # 1列目=time固定、2列目以降がデータ群
        header = rows[0]
        ppm = rows[1]
        data_rows = rows[2:]

        # 列番号 → index化（1始まり）
        col_index = list(range(1, len(header)))

        # ★ number列だけ取り出す（None除去）
        numbers = [h for h in header[1:] if h is not None]

        # number → 列位置辞書
        num_to_idx = {num: i+1 for i, num in enumerate(numbers)}

        # 並び替え（存在するものだけ）
        ordered_numbers = (
            [n for n in custom_order if n in num_to_idx] +
            [n for n in numbers if n not in custom_order]
        )

        # ===== header =====
        ws_out.append(["time"] + ordered_numbers)

        # ===== ppm =====
        ws_out.append(
            ["ppm"] + [ppm[num_to_idx[n]] for n in ordered_numbers]
        )

        # ===== data =====
        for row in data_rows:
            ws_out.append(
                [row[0]] + [row[num_to_idx[n]] for n in ordered_numbers]
            )

    wb_out.save(output_path)
    print("saved:", output_path)


if __name__ == "__main__":

    base = Path(r"C:\Users\haruk\OneDrive - Kyushu Institute Of Technolgy\ドキュメント\lab\program\sekibunti_mottekoi")

    input_file = base / "output" / "result.xlsx"

    reorder_excel_auto(input_file)